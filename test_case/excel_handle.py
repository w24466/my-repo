import openpyxl


class ExcelHandle(object):
    def __init__(self,path,sheet):
        self.path = path
        self.sheet = sheet
    def read_excel(self):
        #打开excel，读取工作簿
        workbook = openpyxl.load_workbook(self.path)
        #指定需要的sheet页
        sheet = workbook[self.sheet]
        #打开所有的sheet
        #sheets=workbook.sheetnames
        #for sheet in sheets:
        #   print sheet
        maxrow = sheet.max_row
        maxcol = sheet.max_column
        line = []
        test_data = []
        # for i in range(2,maxrow+1):
        #     for j in range(1,maxcol+1):
        #         if j == 2 or j ==3:
        #             value = sheet.cell(row=i,column=j).value
        #             line.append(value)
        #             if len(line)==2:
        #                 test_data.append(line)
        for i in range(2,maxrow+1):
            for j in range(1,maxcol+1):
                value = sheet.cell(row=i, column=j).value
                line.append(value)
            test_data.append(line)
            line = []
        return test_data

    def write_back(self,row, rol, result):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb[self.sheet]
        sheet.cell(row, rol).value = result
        wb.save(self.path)  # 保存结果